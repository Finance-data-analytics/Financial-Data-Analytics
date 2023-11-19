import signal
import random
from sqlalchemy import *
import openpyxl
import requests.utils

from main.forms import *
from main import *
from datetime import *
import os
import shutil
from operator import attrgetter
from flask import *
import colorsys

main = Blueprint('main', __name__, template_folder='templates')


@main.before_request
def before_request():
    try:
        db.session.execute('SELECT 1')
        db.session.commit()
    except:
        db.session.rollback()


@application.route('/shutdown', methods=['POST'])
def shutdown():
    with open("server.pid", "r") as f:
        pid = int(f.read())
    try:
        os.kill(pid, signal.SIGINT)  # or SIGTERM
        return "Server shutting down..."
    except Exception as e:
        return str(e)


@main.route("/")
@main.route("/home")
@main.route("/querry_home")
def querry_home_page():
    queries = sql_queries.query.filter_by(group_id=None).all()
    group = group_query.query.all()
    return render_template('home_querry.html', querries=queries, group=group)


COMMON_COLORS = [
    '#F44336',  # Red
    '#E91E63',  # Pink
    '#9C27B0',  # Purple
    '#673AB7',  # Deep Purple
    '#3F51B5',  # Indigo
    '#2196F3',  # Blue
    '#00BCD4',  # Cyan
    '#009688',  # Teal
    '#4CAF50',  # Green
    '#CDDC39',  # Lime
    '#FFEB3B',  # Yellow
    '#FFC107',  # Amber
    '#FF9800',  # Orange
    '#FF5722',  # Deep Orange
    '#795548',  # Brown
    '#607D8B'  # Blue Grey
]


def adjust_luminance(rgb, factor):
    # Convert RGB to HLS
    h, l, s = colorsys.rgb_to_hls(rgb[0] / 255.0, rgb[1] / 255.0, rgb[2] / 255.0)
    # Adjust luminance
    l = max(min(l * factor, 1), 0)
    # Convert HLS back to RGB
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return int(r * 255), int(g * 255), int(b * 255)


def brightness(color):
    r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
    return (r * 299 + g * 587 + b * 114) / 1000  # Luminance formula


def generate_nuances(count):
    chosen_base_color = random.choice(COMMON_COLORS)
    base_rgb = tuple(int(chosen_base_color.lstrip('#')[i:i + 2], 16) for i in (0, 2, 4))
    nuances = set()

    while len(nuances) < count:
        factor = random.uniform(0.5, 1.5)  # Increased the range for a larger gap between nuances
        nuance_rgb = adjust_luminance(base_rgb, factor)
        nuance_hex = '#{:02X}{:02X}{:02X}'.format(*nuance_rgb)
        nuances.add(nuance_hex)

        # Sort the nuances based on brightness in descending order
    sorted_nuances = sorted(nuances, key=brightness, reverse=True)
    return sorted_nuances


@main.route('/jobs')
def jobs():
    queries = {q.Id: q for q in sql_queries.query.all()}
    schedule_querry = {sq.Id: sq for sq in sql_query_executions.query.all()}
    group_dict = {g.id: g for g in group_query.query.all()}
    jobs_data = []
    for job in get_all_jobs():
        job_id_parts = job.id.split("_")

        if job_id_parts[0].isdigit():
            schedule_querry_id = int(job_id_parts[0])
            sq = schedule_querry.get(schedule_querry_id)
            if sq:
                query = queries.get(sq.sql_querry_id)
                group = group_dict.get(sq.group_id)
                if query:
                    jobs_data.append({
                        "job": job,
                        "schedule_querry": sq,
                        "query_name": query.Name,
                        "next_execution": job.next_run_time
                    })
                elif group:
                    # Add logic for handling group queries
                    jobs_data.append({
                        "job": job,
                        "schedule_querry": sq,
                        "query_name": group.group_name,  # use group name as the query name
                        "next_execution": job.next_run_time
                    })

        elif job_id_parts[0] in ['schedule', 'refresh', 'delete']:
            jobs_data.append({
                "job": job,
                "schedule_querry": None,
                "query_name": job.id,
                "next_execution": job.next_run_time
            })

    if jobs_data:
        next_execution_time = min([job["next_execution"] for job in jobs_data])
    else:
        next_execution_time = None

    next_jobs = [job for job in jobs_data if job["next_execution"] == next_execution_time]

    grouped_jobs = {}
    for job in jobs_data:
        query_name = job["query_name"]
        if query_name not in grouped_jobs:
            grouped_jobs[query_name] = []
        grouped_jobs[query_name].append(job)

    nuance_colors = generate_nuances(len(grouped_jobs) + len(next_jobs))
    return render_template('jobs.html', next_jobs=next_jobs, grouped_jobs=grouped_jobs, rdm_color=nuance_colors,
                           length_of_rdm_color=len(nuance_colors))


@main.route("/schedule")
def schedule_querry():
    queries = sql_queries.query.all()
    group = group_query.query.all()
    schedule_querry = sql_query_executions.query.all()
    return render_template('home_schedule.html', queries=queries, schedule_querry=schedule_querry, group=group)


@main.route('/results', methods=['GET', 'POST'])
def Results():
    base_path = os.path.join(".", "Result")
    folders = []
    for folder_name in os.listdir(base_path):
        folder_path = os.path.join(base_path, folder_name)
        if os.path.isdir(folder_path):
            is_empty = len(os.listdir(folder_path)) == 0  # This is a new line.
            folders.append({"name": folder_name, "is_empty": is_empty})
    return render_template('results.html', folders=folders)


@main.route('/results/<folder>')
def results_detail(folder):
    folder_path = os.path.join(".", "Result", folder)
    files = []

    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)

        if filename.endswith(".csv"):
            with open(file_path, 'r', encoding='utf-8') as f:
                line_count = sum(1 for _ in f)  # Count lines in the CSV file
        elif filename.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            line_count = sheet.max_row  # Count rows in the Excel file
        else:
            line_count = None  # Handle other file types as needed

        files.append({
            "name": filename,
            "date_modified": datetime.fromtimestamp(os.path.getmtime(file_path)).strftime("%Y-%m-%d %H:%M:%S"),
            "path": file_path,
            "Zero_row_returned": line_count <= 1 if line_count is not None else False
        })

    return render_template('results_detail.html', files=files)


@main.route('/clean_results', methods=['POST'])
def clean_results():
    base_path = os.path.join(".", "Result")

    query_names_in_db = [query.Name for query in sql_queries.query.all()]
    group_names_in_db = [group.group_name for group in group_query.query.all()]

    for folder_name in os.listdir(base_path):
        folder_path = os.path.join(base_path, folder_name)
        if os.path.isdir(folder_path):
            is_empty = len(os.listdir(folder_path)) == 0
            if is_empty and folder_name not in query_names_in_db and folder_name not in group_names_in_db:
                shutil.rmtree(folder_path)

    return redirect(url_for('main.Results'))


@main.route('/open_file/<path:file_path>')
def open_file(file_path):
    return send_file(file_path, as_attachment=True)


@main.route('/view_file/<path:file_path>')
def view_file(file_path):
    file_extension = os.path.splitext(file_path)[-1]
    sheet_number = request.args.get('sheet_number', 1, type=int)

    if file_extension == '.csv':
        with open(file_path, 'r') as f:
            content = f.read()
            return render_template('view_file.html', content=content, sheet_number=sheet_number, total_sheets=1)

    elif file_extension == '.xlsx':
        workbook = load_workbook(file_path, data_only=True)
        total_sheets = len(workbook.worksheets)
        sheet = workbook.worksheets[sheet_number - 1]  # -1 because list indices start at 0
        content = []
        for row in sheet.iter_rows():
            row_values = [cell.value for cell in row]
            content.append(','.join(map(str, row_values)))
        content_str = '\n'.join(content)
        return render_template('view_file_multi.html', content=content_str, sheet_number=sheet_number, total_sheets=total_sheets)

    else:
        return "Unsupported file format", 400



@main.route("/AddQuerry", methods=['GET', 'POST'])
def add_querry():
    form = AddQuerry()
    if form.validate_on_submit():
        if form.group_name.data:
            group = group_query(group_name=request.form.get('group_name'),
                                Description=description if (description := form.Description.data.strip()) else None,
                                Github_link=github_link if (github_link := form.Github_link.data.strip()) else None)
            db.session.add(group)
            db.session.commit()
            group_id = group.id
        else:
            group_id = None

        dir_name = form.Name.data.strip()
        dir_path = os.path.join(".", "Result", dir_name)
        os.makedirs(dir_path, exist_ok=True)

        # Determine if queries belong to a group
        is_grouped = bool(group_id)
        suffix = "_grouped" if is_grouped else ""
        query_ids = []  # List to store the IDs of added queries
        # Save primary query
        all_queries = [request.form.get(key) for key in request.form if key.startswith('Querry-')]
        for query_text in all_queries:
            new_query = sql_queries(Name=form.Name.data.strip() + suffix,
                                    Querry=query_text,
                                    Description=description if (description := form.Description.data.strip()) else None,
                                    Github_link=github_link if (github_link := form.Github_link.data.strip()) else None,
                                    group_id=group_id)
            db.session.add(new_query)
            db.session.commit()
            query_ids.append(new_query.Id)  # Collecting the ID

        if group_id is not None:
            # Fetch the relevant group_query row
            group_row = group_query.query.get(group_id)

            # If query_ids is a list of IDs, simply assign it.
            # Convert it to a string representation for storage
            group_row.queries_id = str(query_ids)

            # Commit the changes
            db.session.commit()

        return redirect(url_for('main.querry_home_page'))

    if form.errors:
        for err_msg in form.errors.values():
            flash(f'There was an error with creating the query: {err_msg}', category='danger')
    return render_template('addquerry.html', form=form)


@main.route("/delete_querry/<int:id>", methods=["POST"])
def delete_querry(id):
    querry = sql_queries.query.get(id)
    if querry is None:
        flash('Querry not found!', category='danger')
        return redirect(url_for('main.querry_home_page'))

    querry_folder_path = os.path.join(".", "Result", querry.Name)

    if os.path.exists(querry_folder_path) and not os.listdir(querry_folder_path):
        os.rmdir(querry_folder_path)

    if querry.group_id:
        group = group_query.query.get(querry.group_id)
        # Convert the string representation back to a list
        query_ids_list = ast.literal_eval(group.queries_id)
        if querry.Id in query_ids_list:
            query_ids_list.remove(querry.Id)

            # Convert the list back to string representation and update the group_query row
            group.queries_id = str(query_ids_list)

            # Commit the changes to the database
            db.session.commit()
        # group = group_query.query.get(querry.group_id)
        # if group.queries_id:
        #     print(group.queries_id)
        #     db.session.delete(group)


    executions = sql_query_executions.query.filter_by(sql_querry_id=id).all()
    if executions:
        for execution in executions:
            db.session.delete(execution)

    db.session.delete(querry)
    db.session.commit()

    flash('Querry and its executions deleted!', category='success')
    if not querry.group_id:
        return redirect(url_for('main.querry_home_page'))
    elif querry.group_id:
        return redirect(url_for('main.group_details',id=querry.group_id))


@main.route("/delete_group/<int:id>", methods=["POST"])
def delete_group(id):
    group = group_query.query.get(id)
    if group is None:
        flash('Group of query not found!', category='danger')
        return redirect(url_for('main.querry_home_page'))

    group_folder_path = os.path.join(".", "Result", group.group_name)

    if os.path.exists(group_folder_path) and not os.listdir(group_folder_path):
        os.rmdir(group_folder_path)

    # Fetch all the related queries using the group_id
    related_querries = sql_queries.query.filter_by(group_id=id).all()
    if related_querries:
        for related_querry in related_querries:
            db.session.delete(related_querry)

    executions = sql_query_executions.query.filter_by(sql_querry_id=id).all()
    if executions:
        for execution in executions:
            db.session.delete(execution)

    db.session.delete(group)
    db.session.commit()

    flash('Querry and its executions deleted!', category='success')
    return redirect(url_for('main.querry_home_page'))


@main.route('/delete_execution/<int:id>', methods=['POST'])
def delete_execution(id):
    execution = sql_query_executions.query.get_or_404(id)
    querry_id = execution.sql_querry_id
    group_id = execution.group_id
    db.session.delete(execution)
    db.session.commit()
    remove_all_jobs()
    scheduler.add_job(id='schedule_queries_now',
                      func=schedule_queries,
                      trigger='date',
                      run_date=datetime.now())

    # Redirect based on the existence of group_id
    if group_id:
        return redirect(url_for('main.group_details', id=group_id))
    else:
        return redirect(url_for('main.querry_details', id=querry_id))


@main.route('/querry_details/<int:id>', methods=['GET'])
def querry_details(id):
    try:
        customers_prod_lu = application.customers_prod_lu
    except AttributeError:
        flash("Customers for LU have not been loaded!", category='warning')
        customers_prod_lu = []  # or whatever default value you want

    try:
        customers_prod_ch = application.customers_prod_ch
    except AttributeError:
        flash("Customers for CH have not been loaded!", category='warning')
        customers_prod_ch = []  # or whatever default value you want

    query = sql_queries.query.get_or_404(id)

    executions = sorted(query.executions, key=attrgetter('Id'))
    multi_intraday_list = []

    for execution in executions:

        # Get the string representation from the database
        multi_intraday_str = execution.Multi_Intraday_schedule

        if multi_intraday_str:
            # Convert the string to a list of lists
            multi_intraday_list = ast.literal_eval(multi_intraday_str)

    return render_template('details_querry.html', querry=query, executions=executions, customerCh=customers_prod_ch,
                           customerLu=customers_prod_lu, multi_intraday=multi_intraday_list)


@main.route('/group_details/<int:id>', methods=['GET'])
def group_details(id):
    try:
        customers_prod_lu = application.customers_prod_lu
    except AttributeError:
        flash("Customers for LU have not been loaded!", category='warning')
        customers_prod_lu = []  # or whatever default value you want

    try:
        customers_prod_ch = application.customers_prod_ch
    except AttributeError:
        flash("Customers for CH have not been loaded!", category='warning')
        customers_prod_ch = []  # or whatever default value you want

    group = group_query.query.get_or_404(id)
    related_queries = sql_queries.query.filter(
        (sql_queries.group_id == id)
    ).all()
    executions = sorted(group.execution_query, key=attrgetter('Id'))
    multi_intraday_list = []

    for execution in executions:

        # Get the string representation from the database
        multi_intraday_str = execution.Multi_Intraday_schedule

        if multi_intraday_str:
            # Convert the string to a list of lists
            multi_intraday_list = ast.literal_eval(multi_intraday_str)

    return render_template('details_group.html', group=group, executions=executions, customerCh=customers_prod_ch,
                           customerLu=customers_prod_lu, multi_intraday=multi_intraday_list,
                           group_query=related_queries)


@main.route('/execute_schedule/<int:id>', methods=['POST'])
def execute_schedule(id):
    execution = sql_query_executions.query.get_or_404(id)
    if execution.group_id:
        Id = execution.group_id
    elif execution.sql_querry_id:
        Id = execution.sql_querry_id
    scheduler.add_job(id='execute_queries_now',
                      func=process_query_execution,
                      args=(id,),
                      trigger='date',
                      run_date=datetime.now())
    if execution.sql_querry_id:
        return redirect(url_for('main.querry_details', id=Id))
    elif execution.group_id:
        return redirect(url_for('main.group_details', id=Id))


@main.route('/update_querry', methods=['POST'])
def update_querry():
    querry_id = request.form.get('querry_id')
    new_querry = request.form.get('new_querry')
    new_name = request.form.get('new_name').strip()
    new_Descritpion = request.form.get('new_Descritpion')
    new_Github_link = request.form.get('new_Github_link')

    querry = sql_queries.query.get_or_404(querry_id)

    old_folder_path = os.path.join(".", "Result", querry.Name)
    new_folder_path = os.path.join(".", "Result", new_name)

    querry.Querry = new_querry.strip()
    querry.Name = new_name
    description = None if (new_Descritpion == "") or (new_Descritpion == "None") else new_Descritpion.strip()
    github_link = None if (new_Github_link == "") or (new_Github_link == "None") else new_Github_link.strip()

    querry.Description = description
    querry.Github_link = github_link
    db.session.commit()

    if os.path.exists(old_folder_path):
        shutil.move(old_folder_path, new_folder_path)
    flash('the querry and the name were changed', category='info')
    return redirect(url_for('main.querry_details', id=querry_id))


@main.route('/update_group', methods=['POST'])
def update_group():
    group_id = request.form.get('group_id')
    new_name = request.form.get('new_name').strip()
    new_Descritpion = request.form.get('new_Descritpion')
    new_Github_link = request.form.get('new_Github_link')

    group = group_query.query.get_or_404(group_id)

    old_folder_path = os.path.join(".", "Result", group.group_name)
    new_folder_path = os.path.join(".", "Result", new_name)

    # Update main query properties
    group.group_name = new_name
    description = None if (new_Descritpion == "") or (new_Descritpion == "None") else new_Descritpion.strip()
    github_link = None if (new_Github_link == "") or (new_Github_link == "None") else new_Github_link.strip()
    group.Description = description
    group.Github_link = github_link

    # Handle updates/addition of subqueries
    for key, value in request.form.items():
        if key.startswith('new_query_'):
            related_query_id_part = key.split('_')[-1]

            if related_query_id_part == "added":
                new_related_query = sql_queries(
                    Name=new_name + "_grouped",
                    Querry=value,
                    group_id=group_id,
                    Description=description,
                    Github_link=github_link
                )
                query.group_id = group_id
                db.session.add(new_related_query)
                db.session.commit()
                # Convert the string representation back to a list
                query_ids_list = ast.literal_eval(group.queries_id)
                query_ids_list.append(new_related_query.Id)
                # Convert the list back to string representation and update the group_query row
                group.queries_id = str(query_ids_list)

    db.session.commit()  # Committing all changes

    # Handling folder move
    if os.path.exists(old_folder_path):
        shutil.move(old_folder_path, new_folder_path)

    flash('The query and the name were changed', category='info')
    return redirect(url_for('main.group_details', id=group_id))


def is_valid_email(email):
    # Regular expression to validate an email
    pattern = re.compile(r"[^@]+@[^@]+\.[^@]+")
    return pattern.match(email)


@main.route("/schedule_Querry/<int:id>", methods=['GET', 'POST'])
@main.route("/schedule_Querry", defaults={'id': None}, methods=['GET', 'POST'])
def add_schedule_querry(id=None):
    try:
        customers_prod_lu = application.customers_prod_lu
    except AttributeError:
        flash("Customers for LU have not been loaded!", category='warning')
        customers_prod_lu = []  # or whatever default value you want

    try:
        customers_prod_ch = application.customers_prod_ch
    except AttributeError:
        flash("Customers for CH have not been loaded!", category='warning')
        customers_prod_ch = []  # or whatever default value you want
    queries = sql_queries.query.filter_by(group_id=None).all()
    group = group_query.query.all()
    form = ScheduleQuerry()
    if form.validate_on_submit():
        multi_intraday_data = form.multi_intraday_data.data
        multi_intraday_boolean = form.Multi_Intraday.data
        if multi_intraday_data == "[]":
            multi_intraday_data = None
        TimeExecution = None
        schedule_Begin = None
        schedule_End = None
        schedule_Frequency = None

        if form.Intraday_Schedule.data:
            schedule_Begin = form.Intraday_schedule_Begin.data
            schedule_End = form.Intraday_schedule_End.data
            schedule_Frequency = form.Intraday_schedule_Frequency.data
            if multi_intraday_data == "[]":
                multi_intraday_data = None
        if form.Is_not_Intraday.data:
            TimeExecution = form.Time_Execution.data
            if multi_intraday_data == "[]":
                multi_intraday_data = None

        if form.have_parameters.data:
            if form.DateBegin.data == "None":
                DateBegin = None
            else:
                DateBegin = form.DateBegin.data
            if form.DateEnd.data == "None":
                DateEnd = None
            else:
                DateEnd = form.DateEnd.data
            if form.selected_customer_ids.data == "None":
                Customer_Id = None
            else:
                Customer_Id = form.selected_customer_ids.data
            if form.Margin.data == "None":
                Margin = None
            else:
                Margin = form.Margin.data

        else:
            DateBegin = None
            DateEnd = None
            Customer_Id = None
            Margin = None

        if form.list_email.data == "":
            list_email = None
        else:
            list_email = form.list_email.data
        if form.list_email_cc.data == "":
            list_cc_email = None
        else:
            list_cc_email = form.list_email_cc.data

        if form.send_mail.data and list_email is None:
            flash(f'You set True to send an email but the list of email is empty', category='danger')
            return render_template('add_schedule_querry.html', form=form, customerCh=customers_prod_ch,
                                   customerLu=customers_prod_lu, querry=queries, selected_id=id)
        if form.send_mail.data:
            # split email list by either comma or semicolon
            email_list = re.split(r'[;,]', list_email)
            # filter out any empty strings
            email_list = [email.strip() for email in email_list if email.strip()]

            # Check if all emails in the list are valid
            if not all(is_valid_email(email) for email in email_list):
                flash('Please enter a valid list of emails separated by , or ;', category='danger')
                return render_template('add_schedule_querry.html', form=form, customerCh=customers_prod_ch,
                                       customerLu=customers_prod_lu, querry=queries, selected_id=id)

        formatted_week_days = ','.join(day.strip() for day in form.weekly_day.data)
        formated_extension_file = ','.join([extension.strip() for extension in form.file_extension.data])

        if form.frequency.data == "daily" and form.start_date.data == form.end_date.data:
            flash('The begin day and End day cannot be the same. Change the days or choose weekly to choose one day.',
                  category='danger')
            return render_template('add_schedule_querry.html', form=form, customerCh=customers_prod_ch,
                                   customerLu=customers_prod_lu, querry=queries, selected_id=id)
        if form.frequency.data == "weekly" and not formatted_week_days:
            flash('Please choose at least one day for the frequency weekly', category='danger')
            return render_template('add_schedule_querry.html', form=form, customerCh=customers_prod_ch,
                                   customerLu=customers_prod_lu, querry=queries, selected_id=id)

        if multi_intraday_boolean and not multi_intraday_data:
            flash('The multi-intraday list is empty', category='danger')
            return render_template('add_schedule_querry.html', form=form, customerCh=customers_prod_ch,
                                   customerLu=customers_prod_lu, querry=queries, selected_id=id)

        if form.Is_not_Intraday.data and not form.Time_Execution.data:
            flash('Time Execution cannot be empty when "Not Intraday" is selected.', category='danger')
            return render_template('add_schedule_querry.html', form=form, customerCh=customers_prod_ch,
                                   customerLu=customers_prod_lu, querry=queries, selected_id=id)

        if form.Intraday_Schedule.data and (
                not form.Intraday_schedule_Begin.data or not form.Intraday_schedule_End.data or not form.Intraday_schedule_Frequency.data):
            flash('Intraday Begin, End and Frequency cannot be empty when "Intraday Schedule" is selected.',
                  category='danger')
            return render_template('add_schedule_querry.html', form=form, customerCh=customers_prod_ch,
                                   customerLu=customers_prod_lu, querry=queries, selected_id=id)
        if not formated_extension_file and (form.send_slack or form.send_mail):
            flash('You must choose an extension for your output file',
                  category='danger')
            return render_template('add_schedule_querry.html', form=form, customerCh=customers_prod_ch,
                                   customerLu=customers_prod_lu, querry=queries, selected_id=id)
        value = form.selected_query_ids.data
        id, obj_type = value.split(',')
        int(id)
        if obj_type =="query":
            sql_querry_id=id
            group_id=None
        elif obj_type =="group":
            sql_querry_id = None
            group_id = id
        schedule_to_create = sql_query_executions(
            Name=form.Name.data,
            sql_querry_id=sql_querry_id,
            DateBegin=DateBegin,
            DateEnd=DateEnd,
            Customer_Id=Customer_Id,
            Margin=Margin,
            envmt_Test=form.envmt_Test.data,
            envmt_Prod_lu=form.envmt_Prod_lu.data,
            envmt_Prod_ch=form.envmt_Prod_ch.data,
            frequency=form.frequency.data,
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            weekly_day=formatted_week_days,
            Intraday_Schedule=form.Intraday_Schedule.data,
            Intraday_schedule_Begin=schedule_Begin,
            Intraday_schedule_End=schedule_End,
            Intraday_schedule_Frequency=schedule_Frequency,
            Time_Execution=TimeExecution,
            Is_not_Intraday=form.Is_not_Intraday.data,
            have_parameters=form.have_parameters.data,
            Send_slack_empty=form.Send_slack_empty.data,
            Send_mail_empty=form.Send_mail_empty.data,
            list_email=list_email,
            list_cc_email=list_cc_email,
            Multi_Intraday_schedule=multi_intraday_data,
            Multi_Intraday=multi_intraday_boolean,
            send_to_slack=form.send_slack.data,
            send_mail=form.send_mail.data,
            file_extension=formated_extension_file,
            group_id=group_id
        )

        db.session.add(schedule_to_create)
        db.session.commit()
        scheduler.add_job(id='schedule_queries_now',
                          func=schedule_queries,
                          trigger='date',
                          run_date=datetime.now())

        if obj_type == "group":
            return redirect(url_for('main.group_details',id=id))
        elif obj_type == "query":
            return redirect(url_for('main.querry_details',id=id))

    if form.errors != {}:
        for err_msg in form.errors.values():
            flash(f'There was an error scheduling the query: {err_msg}', category='danger')
    return render_template('add_schedule_querry.html', form=form, customerCh=customers_prod_ch,
                           customerLu=customers_prod_lu, query=queries, selected_id=id, group=group)


def has_overlapping_timeslots(multi_intraday_list):
    # Sort the list based on start times
    sorted_list = sorted(multi_intraday_list, key=lambda x: x[0])

    for i in range(1, len(sorted_list)):
        # If the end time of the previous slot is strictly greater than the start time of the current slot
        if sorted_list[i - 1][1] > sorted_list[i][0]:
            return True

    return False


@main.route('/update_execution', methods=['POST'])
def update_execution():
    new_name_execution = request.form.get('nameExecution')
    new_margin = request.form.get('new_margin')
    querry_id = request.form.get('querry_id')

    execution_id = request.form.get('execution_id')
    new_envmt_test = request.form.get('new_envmt_test') == 'true'
    new_envmt_prod_lu = request.form.get('new_envmt_prod_lu') == 'true'
    new_envmt_prod_ch = request.form.get('new_envmt_prod_ch') == 'true'
    new_start_date = request.form.get('new_start_date')
    new_end_date = request.form.get('new_end_date')
    new_frequency = request.form.get('new_frequency')
    new_status = request.form.get('new_status') == 'true'
    new_customer = request.form.get('new_customer')

    newquerry_begin_date = request.form.get('newquerry_begin_date')
    newquerry_end_date = request.form.get('newquerry_end_date')

    new_time_execution = request.form.get('new_time_execution')
    new_is_not_intraday_schedule = request.form.get('new_is_not_intraday_schedule') == 'true'
    new_intraday_schedule = request.form.get('new_intraday_schedule') == 'true'
    new_intraday_begin = request.form.get('new_intraday_begin')
    new_intraday_end = request.form.get('new_intraday_end')
    new_intraday_frequency = request.form.get('new_intraday_frequency').replace(',', '.')
    new_intraday_weekday = request.form.get('new_intraday_weekday')
    have_parameters = request.form.get('have_parameters') == 'true'

    send_slack_empty = request.form.get('new_send_slack_empty') == 'true'
    send_mail_empty = request.form.get('new_send_mail_empty') == 'true'
    list_email = request.form.get('new_listemail')
    list_email_cc = request.form.get('new_listemail_cc')
    Multi_intraday = request.form.get('new_Multi_Intraday') == 'true'
    list_multi_intraday = request.form.get('new_Multi_Intraday_List')
    send_slack = request.form.get('new_send_slack') == 'true'
    send_mail = request.form.get('new_send_mail') == 'true'
    file_extension = request.form.get('new_file_extension')

    if new_intraday_begin > new_intraday_end:
        flash('Begining of the schedule should be before the end of it.', category='danger')
        return redirect(url_for('main.querry_details', id=querry_id))
    # if not new_margin.replace('.', '', 1).isdigit():
    # flash('Only numbers are allowed for the margin.', category='danger')
    # return redirect(url_for('main.querry_details', id=querry_id))

    else:

        execution = sql_query_executions.query.get_or_404(execution_id)
        execution.have_parameters = have_parameters
        if have_parameters:
            if new_customer == "None" or new_customer == "":
                execution.Customer_Id = None
            else:
                execution.Customer_Id = new_customer
            if newquerry_begin_date == "None" or newquerry_begin_date == "":
                execution.DateBegin = None
            else:
                execution.DateBegin = newquerry_begin_date
            if new_margin == "None" or new_margin == "":
                execution.Margin = None
            else:
                execution.Margin = new_margin

            if new_margin == "0.0":
                execution.Margin = None
            if newquerry_end_date == "None" or newquerry_end_date == "":
                execution.DateEnd = None
            else:
                execution.DateEnd = newquerry_end_date

        if not have_parameters:
            execution.Margin = None
            execution.DateBegin = None
            execution.DateEnd = None
            execution.Customer_Id = None

        if new_intraday_schedule:
            execution.Time_Execution = None
            execution.Intraday_schedule_Begin = new_intraday_begin
            execution.Intraday_schedule_End = new_intraday_end
            execution.Intraday_schedule_Frequency = new_intraday_frequency

        if new_is_not_intraday_schedule:
            execution.Time_Execution = new_time_execution
            execution.Intraday_schedule_Begin = None
            execution.Intraday_schedule_End = None
            execution.Intraday_schedule_Frequency = None

        if new_frequency == 'daily':
            execution.weekly_day = None
        if new_frequency == 'weekly':
            execution.weekly_day = ','.join([day.strip() for day in new_intraday_weekday.split(',')])

        execution.Name = new_name_execution
        execution.Intraday_Schedule = new_intraday_schedule
        execution.Is_not_Intraday = new_is_not_intraday_schedule

        execution.envmt_Test = new_envmt_test
        execution.envmt_Prod_lu = new_envmt_prod_lu
        execution.envmt_Prod_ch = new_envmt_prod_ch
        execution.start_date = new_start_date
        execution.end_date = new_end_date
        execution.frequency = new_frequency
        execution.Status = new_status
        execution.file_extension = ','.join([extension.strip() for extension in file_extension.split(',')])
        execution.Send_slack_empty = send_slack_empty
        execution.Send_mail_empty = send_mail_empty
        execution.Multi_Intraday = Multi_intraday

        execution.send_to_slack = send_slack
        execution.send_mail = send_mail

        # Check if the string is not empty or None
        if list_multi_intraday and list_multi_intraday != "[]":
            list_multi_intraday_str = ast.literal_eval(list_multi_intraday)

            if has_overlapping_timeslots(list_multi_intraday_str):
                flash('There are overlapping time slots in your intraday schedule please check!', category='danger')
                return redirect(url_for('main.querry_details', id=querry_id))
            else:
                execution.Multi_Intraday_schedule = list_multi_intraday
        else:
            execution.Multi_Intraday_schedule = None

        if list_email == "" or list_email == "None":
            execution.list_email = None
        else:
            execution.list_email = list_email
        if list_email_cc == "" or list_email_cc == "None":
            execution.list_cc_email = None
        else:
            execution.list_cc_email = list_email_cc

        if execution.list_email is None and execution.send_mail:
            flash(f'You set True to send an email but the list of email is empty', category='danger')
            return redirect(url_for('main.querry_details', id=querry_id))

        db.session.commit()
        scheduler.add_job(id='schedule_queries_now',
                          func=schedule_queries,
                          trigger='date',
                          run_date=datetime.now())
    flash('Execution updated successfully!', category='success')
    return redirect(url_for('main.querry_details', id=querry_id))
