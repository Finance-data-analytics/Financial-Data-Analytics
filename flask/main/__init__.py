import ast
import csv
import re
import smtplib
import traceback
from email import encoders
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from logging.handlers import *
import requests
from flask import *
from datetime import *

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from pytz import *
import logging
from main.extensions import *
from main.models import *
from config import *
import paramiko
import os
import queue
import threading
import os.path
from email.mime.text import MIMEText
import time
import xlsxwriter
from openpyxl import load_workbook
import logging
import traceback

application = Flask(__name__)


class CustomTimedRotatingFileHandler(TimedRotatingFileHandler):
    def __init__(self, dir, when='midnight', interval=1, backupCount=0):
        self.dir = dir
        self.prefix = 'app'
        self.suffix = '%Y-%m-%d.log'
        self.extMatch = re.compile(r"^%s\.\d{4}-\d{2}-\d{2}\.log$" % self.prefix)
        TimedRotatingFileHandler.__init__(self, self.calculateFileName(), when, interval, backupCount)

    def calculateFileName(self):
        return "%s/%s-%s" % (self.dir, self.prefix, datetime.now().strftime(self.suffix))

    def doRollover(self):
        # Construct the new log file name using the date at the moment of rollover
        new_log_file = self.calculateFileName()

        # Close the existing log file
        if self.stream:
            self.stream.close()
            self.stream = None

        # Rename the current log file to the new name
        if os.path.exists(self.baseFilename):
            os.rename(self.baseFilename, new_log_file)

        # Update the base filename to the new file
        self.baseFilename = new_log_file

        # Open the stream to the new log file
        self.stream = self._open()

        # If there's a backup count, handle removing old log files
        if self.backupCount > 0:
            for s in self.getFilesToDelete():
                os.remove(s)


tz = timezone('Europe/Paris')


def time_to_seconds(time_obj):
    return (time_obj.hour * 60 + time_obj.minute) * 60 + time_obj.second


def get_file_extension(file_path):
    return os.path.splitext(file_path)[1]


def send_result_to_slack(file_path, envmt_names, queryname, timestamp, query_execution, nblines, noselect):
    try:
        logging.info(f"{queryname} is being sended to slack")
        headers = {
            "Authorization": Config.SLACK_TOKEN_AUTH  # Replace with your actual token
        }

        if noselect:
            payload = {
                "token": Config.SLACK_TOKEN,
                "channel": Config.SLACK_CHANNEL_ID,
                "text": f"The query '{queryname}' was executed successfully."}
            requests.post("https://slack.com/api/files.upload", headers=headers, data=payload)
        else:
            if len(envmt_names) > 1:
                envmt_display = "PROD"
            else:
                envmt_display = envmt_names[0]['name']

            # Assume nblines has only one element if there's one environment, otherwise sum all elements for multiple envmts.
            total_rows = sum(nblines) if len(envmt_names) > 1 else nblines[0]

            initial_comment = f"{queryname} [{envmt_display}]\n{total_rows} row(s) returned"
            file_ext = get_file_extension(file_path)
            file = {'file': open(file_path, 'rb')}
            payload = {
                "filename": f"{query_execution.Name}_{timestamp}{file_ext}",
                "token": Config.SLACK_TOKEN,
                "channels": Config.SLACK_CHANNEL_ID,
                "initial_comment": initial_comment,
            }
            requests.post("https://slack.com/api/files.upload", headers=headers, data=payload, files=file)
            file['file'].close()

    except Exception as e:
        logging.error(f"An error occurred for {queryname}:{e} \n{traceback.format_exc()}")


def get_customers():
    with application.app_context():
        QUERY = "SELECT id from customer where isactive=True AND id not in ('DEMO','PRAXIS');"

        execution_customers = sql_query_executions(
            envmt_Prod_lu=True,
            envmt_Prod_ch=True,
        )

        # Fetch the environments using get_environments function
        envmts = get_environments(execution_customers)

        # Lists to store customer IDs
        customers_prod_lu = []
        customers_prod_ch = []

        for envmt_details in envmts:
            envmt_name = envmt_details['name']
            try:
                # Execute the query for the current environment
                result = execute_query_for_envmt(envmt_details, QUERY)

                # Extract the customer IDs from the result
                customer_ids = save_customer(result)
                # Store the customer IDs in appropriate lists based on the environment
                if envmt_name == 'PROD LU':
                    customers_prod_lu.extend(customer_ids)
                elif envmt_name == 'PROD CH':
                    customers_prod_ch.extend(customer_ids)

            except Exception as e:
                logging.error(f"An error occurred for {envmt_name}: \n{traceback.format_exc()}")

        application.customers_prod_lu = customers_prod_lu
        application.customers_prod_ch = customers_prod_ch

        return customers_prod_lu, customers_prod_ch


def contains_dangerous_operations(query_str):
    """Check if the query contains DELETE or UPDATE operations."""
    query_str = query_str.upper().strip()  # Convert to uppercase and remove leading/trailing spaces
    return query_str.startswith("DELETE") or query_str.startswith("UPDATE")


def process_query_execution(query_execution_id):
    try:
        with application.app_context():
            # Fetch the query_execution from the database
            query_execution = sql_query_executions.query.get(query_execution_id)

            # Define a helper function to process individual queries
            def process_individual_query(query_object, envmts):
                # Logic to execute an individual query
                logging.info(f"Starting execution... {query_object.Name}")
                formatted_query = format_query(query_object, query_execution)
                all_results = []  # This will hold results from all environments
                for envmt in envmts:
                    result = execute_query_for_envmt(envmt, formatted_query)
                    all_results.append(result)
                return all_results

            if query_execution.group_id:
                group = group_query.query.get(query_execution.group_id)
                logging.info(f"Starting group execution... {group.group_name}")

                # Get the environments related to the group
                group_envmts = get_environments(query_execution)

                # Fetch all the query IDs associated with the group
                query_ids = ast.literal_eval(group.queries_id)
                all_group_results = {}
                for q_id in query_ids:
                    individual_query = sql_queries.query.get(q_id)
                    results = process_individual_query(individual_query, group_envmts)
                    all_group_results[individual_query.Name] = results

                # Now, save all results to separate sheets in the same file
                save_group_results_to_csv(all_group_results, query_execution,group, group_envmts, group.group_name,
                                          datetime.now())

            elif query_execution.sql_querry_id:
                individual_query = sql_queries.query.get(query_execution.sql_querry_id)
                envmts = get_environments(query_execution)
                all_results = process_individual_query(individual_query, envmts)
                save_result_csv(all_results, query_execution, individual_query, envmts, individual_query.Name,
                                datetime.now(), False)

    except Exception as e:
        logging.error(f"An error occurred: {e} \n{traceback.format_exc()}")


def adjust_date(date_str, date_type, has_end_date=True):
    # Assuming the format is CURRENT_DAY-x
    offset = int(date_str.split('-')[-1])
    target_date = datetime.today() - timedelta(days=offset)

    # If the date falls on a Saturday
    if target_date.weekday() == 5:
        # If there's both a start date and an end date, move the start date to Thursday.
        if has_end_date and date_type == "DateBegin":
            target_date -= timedelta(days=2)  # Move to Thursday
        else:
            target_date -= timedelta(days=1)  # Move to Friday

    # If the date falls on a Sunday
    elif target_date.weekday() == 6:
        target_date -= timedelta(days=1)  # Move to Friday for all cases

    return target_date


def format_query(query, query_execution):
    if query_execution.have_parameters:
        logging.info("replacement of the parameters...")
        if query_execution.Customer_Id is not None:
            customer_ids = query_execution.Customer_Id.split(', ')
            quoted_customer_ids = [f"'{id}'" for id in customer_ids]
            formatted_customer_ids = ', '.join(quoted_customer_ids)
        else:
            formatted_customer_ids = None

        if query_execution.DateBegin is not None:
            has_end_date = query_execution.DateEnd is not None
            date_begin = f"'{adjust_date(query_execution.DateBegin, 'DateBegin', has_end_date).strftime('%Y-%m-%d')}'"
        else:
            date_begin = query_execution.DateBegin

        if query_execution.DateEnd is not None:
            date_end = f"'{adjust_date(query_execution.DateEnd, 'DateEnd').strftime('%Y-%m-%d')}'"
        else:
            date_end = query_execution.DateEnd

        formatted_query = query.Querry.format(
            Customer_Id=formatted_customer_ids,
            DateBegin=date_begin,
            DateEnd=date_end,
            margin=query_execution.Margin
        )
    else:
        formatted_query = query.Querry

    return formatted_query


def get_environments(query_execution):
    envmt_config = {
        'PROD LU': {'username': r'Apentis\adebaisieux', 'remote_server_ip': '192.168.122.2'},
        'PROD CH': {'username': 'Apentis', 'remote_server_ip': '10.30.7.1'},
        'TEST': {'username': r'Apentis\adebaisieux', 'remote_server_ip': '192.168.122.4'}
    }

    envmts = []

    if query_execution.envmt_Prod_lu:
        envmts.append({'name': 'PROD LU', **envmt_config['PROD LU']})
    if query_execution.envmt_Prod_ch:
        envmts.append({'name': 'PROD CH', **envmt_config['PROD CH']})
    if query_execution.envmt_Test:
        envmts.append({'name': 'TEST', **envmt_config['TEST']})

    return envmts


def execute_query_for_envmt(envmt, formatted_query):
    # Create SSH client
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    decryption_key = Config.decryption_keys.get(
        envmt['name'])  # Assuming envmt has a 'name' field or use appropriate key
    try:
        # Connect to the remote server using private key
        query_with_key = formatted_query.replace('#cle#', f"'{decryption_key}'")
        private_key = paramiko.RSAKey.from_private_key_file(Config.private_key_path)
        ssh.connect(envmt['remote_server_ip'], username=envmt['username'], pkey=private_key)

        # Create SFTP client
        sftp = ssh.open_sftp()

        # Write the formatted query to a file and upload it to the remote server
        with open('query_tool_script.sql', 'w') as f:
            f.write(query_with_key)
            f.flush()
            os.fsync(f.fileno())  # Ensure flush to disk

        sftp.put('query_tool_script.sql', Config.file_path_on_server)

        # Execute the query on the remote server
        stdin, stdout, stderr = ssh.exec_command(Config.command)

        # Read the standard output and error
        stderr_str = stderr.read().decode()
        # Delete the file on the remote server
        sftp.remove(Config.file_path_on_server)

        return stderr_str

    except Exception as e:
        logging.error(f"An error occurred during query execution for {envmt}: {e} \n{traceback.format_exc()}")
        return None

    finally:
        # Clean up
        sftp.close()
        ssh.close()
        os.remove('query_tool_script.sql')


job_queue = queue.Queue()


def execute_query(query_execution_id):
    job_queue.put(query_execution_id)


def worker():
    while True:
        # Get a job from the queue
        query_execution_id = job_queue.get()

        if query_execution_id is None:
            # If the sentinel value (None) is encountered, break the loop
            break

        # Execute the job
        process_query_execution(query_execution_id)

        # Mark the task as done
        job_queue.task_done()


for _ in range(1):
    threading.Thread(target=worker).start()


def merge_rows(rows):
    final_rows = []
    i = 0
    while i < len(rows):
        merged_row = rows[i].rstrip()
        while i + 1 < len(rows) and ('+|' in merged_row or merged_row.endswith('+')):
            parts_current = merged_row.split('|')
            parts_next = rows[i + 1].rstrip().split('|')

            for j in range(len(parts_current)):
                if parts_current[j].endswith('+'):
                    parts_current[j] = parts_current[j].rstrip('+').strip() + ' ' + parts_next[j].strip()

            merged_row = '|'.join(parts_current)
            i += 1

        final_rows.append(merged_row)
        i += 1
    return final_rows


def send_result_email(file_paths, list_email, executionName, cc_email=None):
    email_content = \
        f"""
            Hello,<br><br>
            Please find attached the report generated: {executionName}<br><br>
            Best regards,<br><br>
            <b><span style="color:#ffa41f;">The Apentis team.</span></b><br>
            <div style="display:flex;padding-top:5px;border-top:1px solid #aaaaaa">
                <div style="width:auto">
                    <img src="cid:myimage" alt="Apentis Logo" style="margin-right: 15px;">
                </div>
                <div style="color:#aaaaaa;">
                    This is an automated message. Please do not reply.<br>
                    Apentis S.à.r.l. located in 2a rue Jean Origer, L-2269 Luxembourg<br>
                    +352 263 635 | support@apentis.eu
                </div>
            </div>
        """

    email = Config.email
    password = Config.password

    # Create a multipart message and set headers
    message = MIMEMultipart()
    message['From'] = email
    message['To'] = list_email.replace(';', ',')
    if cc_email and cc_email.strip():
        message['CC'] = cc_email.replace(';', ',')
    message['Subject'] = f'Scheduled report generated: {executionName}'
    message.attach(MIMEText(email_content, 'html'))

    # Attach the logo as an inline image
    LOGO_PATH = r"/Users/hugolemonnier/Desktop/ESIGELEC/APENTIS/test/Sans titre/logo_apentis.png"
    with open(LOGO_PATH, "rb") as img_file:
        img_data = img_file.read()
        img = MIMEImage(img_data, name="logo_apentis.png")
        img.add_header('Content-ID', '<myimage>')
        message.attach(img)

    # Attach the files
    for file_path in file_paths:
        # Detect the file extension
        file_extension = os.path.splitext(file_path)[-1]
        if file_path == LOGO_PATH:
            continue
        with open(file_path, "rb") as attachment:
            if file_extension == ".csv":
                part = MIMEText(attachment.read().decode('utf-8'), "csv", 'utf-8')
            elif file_extension == ".xlsx":
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
            else:
                logging.error(f"Unsupported file type: {file_extension}")
                continue

            filename_only = os.path.basename(file_path)
            part.add_header("Content-Disposition", f'attachment; filename="{filename_only}"')
            message.attach(part)

    # Send the email
    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(email, password)
            recipients = list_email.split(';')
            if cc_email:
                recipients += cc_email.split(';')
            server.sendmail(email, recipients, message.as_bytes())

        logging.info(f"Message sent to {list_email} with CC to {cc_email}")
    except Exception as error:
        logging.error(f"An error occurred {error}: \n{traceback.format_exc()}")


def save_customer(stderr_str):
    result_start_index = stderr_str.find('===SQL QUERY RESULT===')
    if result_start_index != -1:
        result_start_index += len('===SQL QUERY RESULT===')

        # Find the end index of the header
        header_end_index = stderr_str.find('\n', result_start_index) + 19
        if header_end_index != -1:
            header_end_index += len('\n')
        else:
            logging.warning("Header end not found")
            return

        # Extract the actual result from stderr_str
        result_str = stderr_str[header_end_index:]
        # Parse the result_str to get rows
        rows = result_str.strip().split('\n')
        customer_ids = [row.strip() for row in rows if not row.strip().startswith(
            '-') and 'rows)' not in row and 'ligne)' not in row and 'lignes)' not in row]
        customer_ids.remove('id')

        return customer_ids


def write_header(all_results, queryname, execution_time, file_exts, individual_query_name=None):
    try:
        paths = []
        path_file = os.path.join(".", "Result", queryname)
        formatted_time = execution_time.strftime('%Y-%m-%d %H:%M:%S')
        execution_time = formatted_time.replace(':', '.')

        for file_ext in file_exts:
            final_filename = f'{queryname}{execution_time}.{file_ext}'
            full_path = os.path.join(path_file, final_filename)
            paths.append(full_path)

            if not os.path.exists(path_file):
                os.makedirs(path_file)

            first_result = all_results[0]
            header_start_index = first_result.find('===SQL QUERY RESULT===')

            if header_start_index != -1:
                header_start_index += len('===SQL QUERY RESULT===') + 20
                header_end_index = first_result.find('\n', header_start_index)
                if header_end_index != -1:
                    header_str = first_result[header_start_index:header_end_index].strip()
                    header_cells = [cell.strip() for cell in header_str.split('|')]

                    if file_ext == 'xlsx':
                        if os.path.exists(full_path):
                            # Load existing workbook using openpyxl
                            workbook = load_workbook(full_path)
                        else:
                            # Create new workbook using openpyxl
                            workbook = Workbook()
                            # remove the default sheet created by openpyxl
                            default_sheet = workbook.active
                            workbook.remove(default_sheet)

                        # Create new worksheet with the name of the individual query
                        worksheet_name = individual_query_name if individual_query_name else 'Sheet1'
                        if worksheet_name not in workbook.sheetnames:
                            worksheet = workbook.create_sheet(title=worksheet_name)
                        else:
                            worksheet = workbook[worksheet_name]

                        for j, cell in enumerate(header_cells):
                            worksheet.cell(row=1, column=j + 1, value=cell)
                        workbook.save(full_path)
                    if file_ext == 'csv':
                        with open(full_path, 'w', newline='', encoding='utf-8-sig') as csv_file:
                            csv_writer = csv.writer(csv_file, quotechar='"', quoting=csv.QUOTE_ALL)
                            csv_writer.writerow(header_cells)

                else:
                    logging.error(first_result)
                    logging.warning("Header end not found")
        return paths
    except Exception as e:
        logging.error(f"An error occurred: {e} \n{traceback.format_exc()}")


def save_result_csv(all_results, query_execution, query, envmts, queryname, execution_time, group):
    nb_line = []
    try:
        desired_extensions = query_execution.file_extension.split(',')
        paths = write_header(all_results, queryname, execution_time, desired_extensions)
        for file_ext, full_path in zip(desired_extensions, paths):
            for i, stderr_str in enumerate(all_results):
                # Find the start index of the result in stderr_str
                result_start_index = stderr_str.find('===SQL QUERY RESULT===')
                if result_start_index != -1:
                    result_start_index += len('===SQL QUERY RESULT===')

                    # Find the end index of the header
                    header_end_index = stderr_str.find('\n', result_start_index) + 19
                    if header_end_index != -1:
                        header_end_index += len('\n')
                    else:
                        logging.warning("Header end not found")
                        return

                    # Extract the actual result from stderr_str
                    result_str = stderr_str[header_end_index:]
                    # Parse the result_str to get rows
                    rows = result_str.strip().split('\n')
                    # Process and merge the rows
                    merged_rows = merge_rows(rows)

                    # Remove unwanted rows
                    del merged_rows[0:2]
                    del merged_rows[-1]
                    # Prepare the rows for CSV writing
                    csv_rows = []
                    for row in merged_rows:
                        modified_row = []
                        for cell in row.split('|'):
                            stripped_cell = cell.strip().replace('\n', ' ')
                            if stripped_cell == 't':
                                modified_row.append('True')
                            elif stripped_cell == 'f':
                                modified_row.append('False')
                            elif stripped_cell == 'D':
                                modified_row.append('Daily')
                            elif stripped_cell == 'W':
                                modified_row.append('Weekly')
                            else:
                                modified_row.append(stripped_cell)
                        csv_rows.append(modified_row)
                    if file_ext == 'csv':
                        with open(full_path, 'a', newline='', encoding='utf-8-sig') as final_csv:
                            csv_writer = csv.writer(final_csv, quotechar='"', quoting=csv.QUOTE_ALL)
                            csv_writer.writerows(csv_rows)  # Write the data rows

                    elif file_ext == 'xlsx':
                        workbook = load_workbook(full_path)
                        sheet = workbook.active
                        max_row = sheet.max_row + 1
                        for row in csv_rows:
                            for col_num, cell_value in enumerate(row, 1):
                                sheet.cell(row=max_row, column=col_num, value=cell_value)
                            max_row += 1

                            # Finally, save the workbook
                        workbook.save(full_path)
                    nblines = len(csv_rows)
                    nb_line.append(nblines)

                else:
                    logging.info("Header not found")
                    return

            logging.info(f"Setting Last_result for query_execution ID {query_execution.Id} to {full_path}")

            query.Last_result = full_path

            # Commit the changes
            db.session.commit()

            # After committing, log the success:
            logging.info(f"Successfully saved Last_result for query_execution ID {query_execution.Id}")

            if query_execution.send_to_slack:
                if all(n == 0 for n in nb_line) and not query_execution.Send_slack_empty:
                    logging.warning(f"the query {queryname} executed at {execution_time} returned no rows")
                if all(n == 0 for n in nb_line) and query_execution.Send_slack_empty or any(n > 0 for n in nb_line):
                    send_result_to_slack(full_path, envmts, queryname, execution_time, query_execution, nb_line, False)

        if query_execution.send_mail:
            if (query_execution.list_email != 'None' and query_execution.list_email is not None) and (
                    (all(n == 0 for n in nb_line) and query_execution.Send_mail_empty) or
                    any(n > 0 for n in nb_line)):
                send_result_email(paths, query_execution.list_email, query_execution.Name,
                                  query_execution.list_cc_email)

    except Exception as e:
        logging.error(f"An error occurred: {e} \n{traceback.format_exc()}")


def save_group_results_to_csv(all_group_results, query_execution,group, group_envmts, group_name, execution_time):
    nb_line = []
    try:
        desired_extensions = query_execution.file_extension.split(',')

        for query_name, results in all_group_results.items():
            paths = write_header(results, group_name, execution_time, desired_extensions, query_name)

            for file_ext, full_path in zip(desired_extensions, paths):
                for i, stderr_str in enumerate(results):
                    # Find the start index of the result in stderr_str
                    result_start_index = stderr_str.find('===SQL QUERY RESULT===')
                    if result_start_index != -1:
                        result_start_index += len('===SQL QUERY RESULT===')

                        # Find the end index of the header
                        header_end_index = stderr_str.find('\n', result_start_index) + 19
                        if header_end_index != -1:
                            header_end_index += len('\n')
                        else:
                            logging.warning("Header end not found")
                            return

                        # Extract the actual result from stderr_str
                        result_str = stderr_str[header_end_index:]
                        # Parse the result_str to get rows
                        rows = result_str.strip().split('\n')
                        # Process and merge the rows
                        merged_rows = merge_rows(rows)

                        # Remove unwanted rows
                        del merged_rows[0:2]
                        del merged_rows[-1]

                        # Prepare the rows for CSV writing
                        csv_rows = []
                        for row in merged_rows:
                            modified_row = []
                            for cell in row.split('|'):
                                stripped_cell = cell.strip().replace('\n', ' ')
                                if stripped_cell == 't':
                                    modified_row.append('True')
                                elif stripped_cell == 'f':
                                    modified_row.append('False')
                                elif stripped_cell == 'D':
                                    modified_row.append('Daily')
                                elif stripped_cell == 'W':
                                    modified_row.append('Weekly')
                                else:
                                    modified_row.append(stripped_cell)
                            csv_rows.append(modified_row)

                        if file_ext == 'csv':
                            with open(full_path, 'a', newline='', encoding='utf-8-sig') as final_csv:
                                csv_writer = csv.writer(final_csv, quotechar='"', quoting=csv.QUOTE_ALL)
                                csv_writer.writerows(csv_rows)  # Write the data rows

                        elif file_ext == 'xlsx':
                            workbook = load_workbook(full_path)
                            worksheet_name = query_name
                            sheet = workbook[worksheet_name]
                            max_row = sheet.max_row + 1
                            for row in csv_rows:
                                for col_num, cell_value in enumerate(row, 1):
                                    sheet.cell(row=max_row, column=col_num, value=cell_value)
                                max_row += 1
                            workbook.save(full_path)

                        nblines = len(csv_rows)
                        nb_line.append(nblines)

                    else:
                        logging.info("Header not found")
                        return

            group.Last_result = full_path
            # Commit the changes
            db.session.commit()

        if query_execution.send_to_slack:
            if all(n == 0 for n in nb_line) and not query_execution.Send_slack_empty:
                logging.warning(f"the query {group_name} executed at {execution_time} returned no rows")
            if all(n == 0 for n in nb_line) and query_execution.Send_slack_empty or any(n > 0 for n in nb_line):
                send_result_to_slack(full_path, group_envmts, group_name, execution_time, query_execution, nb_line,
                                     False)

        if query_execution.send_mail:
            if (query_execution.list_email != 'None' and query_execution.list_email is not None) and (
                    (all(n == 0 for n in nb_line) and query_execution.Send_mail_empty) or any(
                n > 0 for n in nb_line)):
                send_result_email(full_path, query_execution.list_email, query_execution.Name,
                                  query_execution.list_cc_email)

    except Exception as e:
        logging.error(f"An error occurred: {e} \n{traceback.format_exc()}")


def schedule_queries():
    logging.info("schedule_queries is running")
    try:
        # Remove all existing jobs first
        # scheduler.remove_all_jobs()

        # Fetch all scheduled query executions from the database
        with application.app_context():
            scheduled_queries = sql_query_executions.query.all()
            for query in scheduled_queries:

                # Skip scheduling if the status is False
                if not query.Status:
                    logging.warning(f"Skipping query id={query.Id} because status is False")
                    continue

                logging.info(f"Scheduling query id={query.Id}")

                # If no valid database_uri is found, skip scheduling
                if not (query.envmt_Test or query.envmt_Prod_lu or query.envmt_Prod_ch):
                    logging.warning(f"Skipping query id={query.Id} because no valid database URI was found")
                    continue

                # For daily schedule
                if query.frequency == 'daily':
                    if query.Intraday_Schedule:
                        handle_intraday_schedule(query)
                    elif query.Is_not_Intraday:
                        handle_non_intraday_schedule(query)
                    elif query.Multi_Intraday:
                        handle_multi_intraday_schedule(query)

                # For weekly schedule
                elif query.frequency == 'weekly':
                    if query.Intraday_Schedule:
                        handle_intraday_schedule(query, weekly=True)
                    elif query.Is_not_Intraday:
                        handle_non_intraday_schedule(query, weekly=True)
                    elif query.Multi_Intraday:
                        handle_multi_intraday_schedule(query, weekly=True)
                else:
                    logging.warning(f"Skipping query id={query.Id} because no valid scheduling option was found")
    except Exception as e:
        logging.error(f"An error occurred: {e} \n{traceback.format_exc()}")


def handle_intraday_schedule(query, weekly=False):
    start_time = query.Intraday_schedule_Begin
    end_time = query.Intraday_schedule_End
    interval = query.Intraday_schedule_Frequency * 3600

    start_time_seconds = time_to_seconds(start_time)
    end_time_seconds = time_to_seconds(end_time)

    interval_count = int((end_time_seconds - start_time_seconds) / interval) + 1

    if not weekly:  # Daily frequency
        days_of_week = get_weekday_string(query.start_date, query.end_date, None)
    else:  # Weekly frequency
        days_of_week = get_weekday_string(None, None, query.weekly_day)

    tz = timezone('Europe/Paris')

    for day_of_week in days_of_week:
        for i in range(interval_count):
            job_id = f"{query.Id}_{day_of_week}_interval_{i}"
            if scheduler.get_job(job_id):
                scheduler.remove_job(job_id)
            # Calculate time for this job
            job_time_seconds = start_time_seconds + i * interval
            job_time_hour = job_time_seconds // 3600
            job_time_minute = (job_time_seconds % 3600) // 60
            job_time_second = (job_time_seconds % 3600) % 60

            scheduler.add_job(
                id=job_id,
                func=execute_query, args=(query.Id,),
                trigger='cron',
                day_of_week=day_of_week,
                hour=int(job_time_hour),
                minute=int(job_time_minute),
                second=int(job_time_second),
                timezone=tz)

            job = scheduler.get_job(id=job_id)

            if job.next_run_time > next_sunday():
                scheduler.remove_job(job_id)
            # else:
            #    print_job_details(job_id)


def handle_multi_intraday_schedule(query, weekly=False):
    multi_intraday_list = ast.literal_eval(query.Multi_Intraday_schedule)
    n = 0
    executed_times = set()  # Set to keep track of executed times
    for schedule in multi_intraday_list:
        n += 1
        start_time = datetime.strptime(schedule[0], '%H:%M').time()
        end_time = datetime.strptime(schedule[1], '%H:%M').time()
        interval = float(schedule[2].replace(',', '.')) * 3600

        start_time_seconds = time_to_seconds(start_time)
        end_time_seconds = time_to_seconds(end_time)

        interval_count = int((end_time_seconds - start_time_seconds) / interval) + 1

        if not weekly:  # Daily frequency
            days_of_week = get_weekday_string(query.start_date, query.end_date, None)
        else:  # Weekly frequency
            days_of_week = get_weekday_string(None, None, query.weekly_day)

        tz = timezone('Europe/Paris')

        for day_of_week in days_of_week:
            for i in range(interval_count):

                # Calculate time for this job
                job_time_seconds = start_time_seconds + i * interval
                job_time_hour = job_time_seconds // 3600
                job_time_minute = (job_time_seconds % 3600) // 60
                job_time_second = (job_time_seconds % 3600) % 60

                if (day_of_week, job_time_hour, job_time_minute, job_time_second) in executed_times:
                    continue  # Skip this iteration

                    # Add the job time to the set
                executed_times.add((day_of_week, job_time_hour, job_time_minute, job_time_second))

                job_id = f"{query.Id}_{day_of_week}_interval_{i}_{n}"

                if scheduler.get_job(job_id):
                    scheduler.remove_job(job_id)

                scheduler.add_job(
                    id=job_id,
                    func=execute_query, args=(query.Id,),
                    trigger='cron',
                    day_of_week=day_of_week,
                    hour=int(job_time_hour),
                    minute=int(job_time_minute),
                    second=int(job_time_second),
                    timezone=tz)

                job = scheduler.get_job(id=job_id)

                if job.next_run_time > next_sunday():
                    scheduler.remove_job(job_id)
                # else:
                #    print_job_details(job_id)


def handle_non_intraday_schedule(query, weekly=False):
    time = query.Time_Execution
    tz = timezone('Europe/Paris')

    if not weekly:  # Daily frequency
        days_of_week = get_weekday_string(query.start_date, query.end_date, None)
    else:  # Weekly frequency
        days_of_week = get_weekday_string(None, None, query.weekly_day)

    for day_of_week in days_of_week:
        job_id = f"{query.Id}_{day_of_week}"
        if scheduler.get_job(job_id):
            scheduler.remove_job(job_id)
        scheduler.add_job(id=job_id,
                          func=execute_query, args=(query.Id,),
                          trigger='cron', day_of_week=day_of_week, hour=time.hour, minute=time.minute,
                          second=time.second, timezone=tz)

        job = scheduler.get_job(id=job_id)

        if job.next_run_time > next_sunday():
            scheduler.remove_job(job_id)
        # else:
        #    print_job_details(job_id)


def next_sunday():
    today = datetime.now(timezone('Europe/Paris'))  # you can use your preferred timezone here
    diff = (6 - today.weekday()) % 7  # 6 corresponds to Sunday
    next_sunday = today + timedelta(days=diff)
    next_sunday = next_sunday.replace(hour=23, minute=59, second=59, microsecond=0)
    return next_sunday


def get_weekday_string(start_date, end_date=None, weekly_days=None):
    weekdays_full = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
    weekdays_short = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']
    weekday_map = dict(zip(weekdays_full, weekdays_short))

    # For daily frequencies
    if not weekly_days:
        start_index = weekdays_short.index(weekday_map[start_date])
        end_index = weekdays_short.index(weekday_map[end_date]) if end_date else start_index
        return [weekdays_short[i % 7] for i in range(start_index, end_index + 1)]

    # For weekly frequencies
    if weekly_days:
        return [weekday_map[day.strip()] for day in weekly_days.split(',')]


def get_all_jobs():
    return scheduler.get_jobs()


# def print_job_details(job_id):
# job = scheduler.get_job(id=job_id)

# print out the job details
# logging.info(f"Next run time of {job_id}: {job.next_run_time}")

def remove_all_jobs():
    excluded_ids = ['schedule_get_customers', 'refresh_schedule', 'delete_old_files']
    for job in scheduler.get_jobs():
        if job.id not in excluded_ids:
            scheduler.remove_job(job.id)


def delete_old_files():
    directory_path_Result = r"/Users/hugolemonnier/Desktop/test/Sans titre/Result"
    directory_path_logs = r"/Users/hugolemonnier/Desktop/test/Sans titre/logs"

    # Get the current time
    current_time = time.time()
    # Define the age of files to be deleted (1 month in seconds)
    one_month_seconds = 30 * 24 * 60 * 60

    def is_file_old(file_path, is_log_file=False):
        """Determine if a file is older than one month."""
        if is_log_file:
            # Extract the date from the filename
            date_match = re.search(r'(\d{4}-\d{2}-\d{2})', file_path)
            if date_match:
                file_date_str = date_match.group(1)
                file_date = datetime.strptime(file_date_str, '%Y-%m-%d')
                file_timestamp = file_date.timestamp()
            else:
                # If no date is found, use the modified time
                file_timestamp = os.path.getmtime(file_path)
        else:
            file_timestamp = os.path.getmtime(file_path)

        return file_timestamp < (current_time - one_month_seconds)

    for directory in [directory_path_Result, directory_path_logs]:
        for dirpath, dirnames, filenames in os.walk(directory):
            for filename in filenames:
                file_path = os.path.join(dirpath, filename)
                is_log_file = directory == directory_path_logs

                if is_file_old(file_path, is_log_file):
                    os.remove(file_path)
                    logging.info(f"Deleted old file: {filename}")


scheduler.add_job(
    id='schedule_get_customers',
    func=get_customers,
    trigger='cron',
    hour=0,
    minute=0
)

scheduler.add_job(id='refresh_schedule', func=schedule_queries, trigger='interval', minutes=30, timezone=tz)
scheduler.add_job(id='delete_old_files', func=delete_old_files, trigger='cron', hour=0, minute=0)


def setup_logger():
    log_filename = 'logs/app.log'
    handler = TimedRotatingFileHandler(log_filename, when='midnight', interval=1, backupCount=0, encoding='utf-8')
    handler.suffix = '%Y-%m-%d.log'
    handler.extMatch = re.compile(r"^\d{4}-\d{2}-\d{2}.log$")
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)

    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    logger.addHandler(handler)


# Call the setup function to initialize the logger
setup_logger()


def create_app(config_class=Config):
    logging.info("Creating app...")

    global application
    application = Flask(__name__)
    application.config.from_object(config_class)

    db.init_app(application)

    # get_customers()

    logging.getLogger('apscheduler').setLevel(logging.WARNING)

    scheduler.init_app(application)

    from main.routes import main as main_blueprint
    application.register_blueprint(main_blueprint)

    with application.app_context():
        db.create_all()
        scheduler.add_job(id='schedule_queries_now',
                          func=schedule_queries,
                          trigger='date',
                          run_date=datetime.now())
        logging.info("schedule starting...")
    scheduler.start()
    print("App up and running")
    return application
